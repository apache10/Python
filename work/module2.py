import openpyxl
import sys


def readExcel(filename):
	wb = openpyxl.load_workbook(filename)
	list1 = wb.get_sheet_names()
	for k in range(0,len(list1)):
		sheet = wb.get_sheet_by_name(list1[k])
		print "\n\n"+list1[k]+ "\n\n"
		for i in range(1,sheet.max_row+1):
			print "\n"
			for j in range(1, sheet.max_column+1):
				sys.stdout.write(str(sheet.cell(row=i,column=j).value) + "\t")
				sys.stdout.flush()
					
